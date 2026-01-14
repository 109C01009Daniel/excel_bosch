#!/usr/bin/env python3
"""
Excel 搜索系統 - MariaDB CLI 工具
"""
import os
import sys
import click
import time
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm

from database_mariadb import DatabaseManager
from file_scanner import FileScanner
from config_mariadb import DB_CONFIG, BATCH_SIZE


# ============================================================================
# 輔助函數
# ============================================================================

def format_size(bytes_size):
    """格式化檔案大小"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if bytes_size < 1024.0:
            return f"{bytes_size:.2f} {unit}"
        bytes_size /= 1024.0


def print_header(text):
    """列印標題"""
    click.echo()
    click.echo("=" * 70)
    click.echo(f"  {text}")
    click.echo("=" * 70)
    click.echo()


def print_success(text):
    """列印成功訊息"""
    click.secho(f"✅ {text}", fg='green')


def print_error(text):
    """列印錯誤訊息"""
    click.secho(f"❌ {text}", fg='red')


def print_info(text):
    """列印資訊"""
    click.secho(f"ℹ️  {text}", fg='blue')


def print_warning(text):
    """列印警告"""
    click.secho(f"⚠️  {text}", fg='yellow')


def read_excel_file(file_path):
    """
    讀取單個 Excel 檔案

    Returns:
        list: 單元格資料列表
    """
    cells_data = []

    try:
        # 打開檔案（唯讀模式）
        workbook = load_workbook(file_path, read_only=True, data_only=True)

        # 遍歷所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 遍歷所有單元格
            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is not None:
                        cells_data.append({
                            'sheet_name': sheet_name,
                            'row': row_idx,
                            'col': col_idx,
                            'location': f"{cell.column_letter}{cell.row}",
                            'value': str(cell.value),
                        })

        workbook.close()
        return cells_data

    except Exception as e:
        raise Exception(f"讀取檔案失敗: {e}")


# ============================================================================
# CLI 命令
# ============================================================================

@click.group()
def cli():
    """📊 Excel 搜索系統 - MariaDB 版本"""
    pass


@cli.command()
@click.argument('path', type=click.Path(exists=True))
@click.option('--recursive/--no-recursive', default=True, help='是否遞迴搜索子目錄')
@click.option('--incremental/--full', default=True, help='增量索引（僅更新變動檔案）或全量索引')
def index(path, recursive, incremental):
    """📥 索引 Excel 檔案"""
    mode_text = "增量索引" if incremental else "全量索引"
    print_header(f"🔍 開始 {mode_text} Excel 檔案 (MariaDB)")

    # 掃描檔案
    scanner = FileScanner()
    files = scanner.scan_directory(path, recursive=recursive, show_progress=True)

    if not files:
        print_warning("未找到任何 Excel 檔案")
        return

    print_success(f"找到 {len(files)} 個 Excel 檔案")
    click.echo()

    # 連接資料庫
    with DatabaseManager() as db:
        if not db.connection:
            print_error("無法連接到 MariaDB 資料庫")
            return

        # 確保資料表存在
        db.create_tables()

        # 逐個處理檔案
        total_cells = 0
        success_count = 0
        failed_count = 0
        new_files = 0
        updated_files = 0
        skipped_files = 0

        with tqdm(files, desc="索引檔案", unit="file") as pbar:
            for file_info in pbar:
                file_path = file_info['file_path']
                file_name = file_info['file_name']
                file_modified = file_info['last_modified']

                try:
                    pbar.set_description(f"處理: {file_name[:30]}")

                    # 檢查檔案是否已存在於資料庫
                    existing_file = db.get_file_by_path(file_path)

                    # 增量模式：判斷是否需要索引
                    if incremental and existing_file:
                        # 比較修改時間
                        if file_modified <= existing_file['last_modified']:
                            # 檔案未變動，跳過
                            skipped_files += 1
                            pbar.set_description(f"⏭️  跳過: {file_name[:30]}")
                            continue
                        else:
                            # 檔案已更新，刪除舊資料
                            updated_files += 1
                            db.delete_file(file_path)  # CASCADE 會自動刪除相關 cells
                            pbar.set_description(f"🔄 更新: {file_name[:30]}")
                    elif not existing_file:
                        # 新檔案
                        new_files += 1
                        pbar.set_description(f"➕ 新增: {file_name[:30]}")
                    else:
                        # 全量模式：刪除舊資料重新索引
                        if existing_file:
                            db.delete_file(file_path)

                    # 讀取檔案內容
                    cells_data = read_excel_file(file_path)

                    if not cells_data:
                        print_warning(f"⚠️  檔案無內容: {file_name}")
                        continue

                    # 新增檔案記錄
                    file_id = db.add_file(
                        file_path=file_path,
                        file_name=file_name,
                        last_modified=file_modified,  # 已經是 datetime 物件
                        file_size=file_info['file_size']
                    )

                    if not file_id:
                        print_error(f"❌ 新增檔案記錄失敗: {file_name}")
                        failed_count += 1
                        continue

                    # 準備批次插入資料
                    batch_data = []
                    for cell in cells_data:
                        value = cell['value']
                        batch_data.append((
                            file_id,
                            cell['sheet_name'],
                            cell['row'],
                            cell['col'],
                            cell['location'],
                            value,
                            value.lower(),  # value_lower
                            False,  # is_merged
                            None    # merged_range
                        ))

                    # 批次插入
                    inserted = 0
                    for i in range(0, len(batch_data), BATCH_SIZE):
                        batch = batch_data[i:i + BATCH_SIZE]
                        inserted += db.add_cells_batch(batch)

                    # 更新檔案的單元格數量
                    db.update_file_cell_count(file_id, len(cells_data))

                    total_cells += len(cells_data)
                    success_count += 1

                except Exception as e:
                    print_error(f"❌ 索引失敗: {file_name}")
                    print_error(f"   錯誤: {str(e)}")
                    failed_count += 1

        # 反向比對：清理已刪除的檔案
        deleted_files = 0
        if incremental:
            click.echo()
            print_info("🔍 檢查已刪除的檔案...")

            # 取得資料庫中該路徑下的所有檔案
            import os
            base_path = os.path.abspath(path)
            db_files = db.get_files_under_path(base_path)

            # 檢查每個資料庫中的檔案是否還存在
            for db_file in db_files:
                file_path = db_file['file_path']
                if not os.path.exists(file_path):
                    # 檔案已被刪除，從資料庫移除
                    if db.delete_file(file_path):
                        deleted_files += 1
                        print_info(f"🗑️  清除: {db_file['file_name']} (已刪除)")

    # 顯示結果
    click.echo()
    click.echo("─" * 70)
    print_success(f"索引完成！成功: {success_count}, 失敗: {failed_count}")

    if incremental:
        print_info(f"➕ 新增檔案: {new_files}")
        print_info(f"🔄 更新檔案: {updated_files}")
        print_info(f"⏭️  跳過檔案: {skipped_files}")
        if deleted_files > 0:
            print_info(f"🗑️  清除檔案: {deleted_files}")

    print_info(f"📊 總共索引 {total_cells:,} 個單元格")
    click.echo("─" * 70)


@cli.command()
@click.argument('keyword')
@click.option('--limit', default=20, help='結果數量限制')
@click.option('--full-row', is_flag=True, help='顯示完整行內容')
def search(keyword, limit, full_row):
    """🔍 搜索 Excel 內容"""
    print_header(f'🔍 搜索: "{keyword}" (MariaDB)')

    # 連接資料庫並搜索
    with DatabaseManager() as db:
        if not db.connection:
            print_error("無法連接到 MariaDB 資料庫")
            return

        start_time = time.time()
        results = db.search(keyword, limit)
        query_time = (time.time() - start_time) * 1000  # 轉換成毫秒

    if not results:
        print_warning("未找到任何結果")
        return

    # 顯示結果
    print_success(f"找到 {len(results)} 個結果")
    click.echo()

    for idx, result in enumerate(results, 1):
        click.echo("─" * 70)
        click.secho(f"結果 {idx}", fg='cyan', bold=True)
        click.echo(f"📄 檔案: {result['file_name']}")
        click.secho(f"📁 路徑: {result['file_path']}", fg='blue')
        click.echo(f"📊 工作表: {result['sheet_name']}")
        click.echo(f"📍 位置: {result['cell_location']} (第{result['row_num']}行, 第{result['col_num']}列)")

        value = result['value']
        if full_row:
            click.echo(f"📝 內容: {value}")
        else:
            # 截斷顯示
            max_len = 200
            if len(value) > max_len:
                value = value[:max_len] + "..."
            click.echo(f"📝 內容: {value}")

    # 顯示統計
    click.echo()
    click.echo("─" * 70)
    print_info(f"查詢時間: {query_time:.2f} ms")

    if len(results) >= limit:
        print_info(f"僅顯示前 {limit} 個結果，使用 --limit 參數顯示更多")


@cli.command()
def stats():
    """📊 顯示資料庫統計資訊"""
    print_header("📊 資料庫統計 (MariaDB)")

    with DatabaseManager() as db:
        if not db.connection:
            print_error("無法連接到 MariaDB 資料庫")
            return

        stats = db.get_stats()

    if not stats:
        print_error("無法獲取統計資訊")
        return

    # 顯示基本資訊
    click.echo(f"📁 檔案總數: {stats['file_count']:,}")
    click.echo(f"📊 單元格總數: {stats['cell_count']:,}")
    click.echo(f"💾 資料庫大小: {stats['db_size_mb']:.2f} MB")
    click.echo(f"🔗 資料庫: {DB_CONFIG['database']}")
    click.echo(f"🖥️  主機: {DB_CONFIG['host']}:{DB_CONFIG['port']}")
    click.echo()

    # 顯示最近索引的檔案
    if stats['recent_files']:
        click.secho("最近索引的檔案 (前 10 個):", fg='cyan', bold=True)
        click.echo()

        recent_files = stats['recent_files']
        for i, file_info in enumerate(recent_files, 1):
            click.echo(f"  {i}. {file_info['file_name']}")
            click.secho(f"     📍 路徑: {file_info['file_path']}", fg='cyan')
            click.echo(f"     📊 單元格: {file_info['cell_count']} | 索引時間: {file_info['indexed_at']}")
            if i < len(recent_files):
                click.echo()


@cli.command()
@click.confirmation_option(prompt='確定要清空整個資料庫嗎？')
def clear():
    """🗑️  清空資料庫"""
    print_header("🗑️  清空資料庫 (MariaDB)")

    with DatabaseManager() as db:
        if not db.connection:
            print_error("無法連接到 MariaDB 資料庫")
            return

        if db.clear_database():
            print_success("資料庫已清空")
        else:
            print_error("清空資料庫失敗")


@cli.command()
def info():
    """ℹ️  顯示系統資訊"""
    print_header("ℹ️  系統資訊 (MariaDB)")

    click.echo("📦 Excel 搜索系統 - MariaDB 版本")
    click.echo("🏗️  資料庫: MariaDB/MySQL")
    click.echo(f"🔗 連接: {DB_CONFIG['host']}:{DB_CONFIG['port']}")
    click.echo(f"💾 資料庫名稱: {DB_CONFIG['database']}")
    click.echo(f"👤 使用者: {DB_CONFIG['user']}")
    click.echo()
    click.echo("功能:")
    click.echo("  • index <路徑>    - 索引 Excel 檔案")
    click.echo("  • search <關鍵詞> - 搜索內容")
    click.echo("  • stats           - 顯示統計資訊")
    click.echo("  • clear           - 清空資料庫")
    click.echo("  • info            - 顯示系統資訊")
    click.echo()
    click.echo("範例:")
    click.echo("  python3 excel_search_cli_mariadb.py index ./Sharepoint")
    click.echo("  python3 excel_search_cli_mariadb.py search 'IR LED'")
    click.echo("  python3 excel_search_cli_mariadb.py stats")


if __name__ == '__main__':
    cli()
