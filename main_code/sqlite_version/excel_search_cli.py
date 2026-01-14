#!/usr/bin/env python3
"""
Excel 搜索系統 - 簡單 CLI 工具
"""
import os
import sys
import click
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm

from database import Database
from file_scanner import FileScanner
from config import DATABASE_PATH


# ============================================================================
# 輔助函數
# ============================================================================

def get_db():
    """獲取資料庫連接"""
    return Database(DATABASE_PATH)


def format_size(bytes_size):
    """格式化檔案大小"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if bytes_size < 1024.0:
            return f"{bytes_size:.2f} {unit}"
        bytes_size /= 1024.0


def print_header(text):
    """列印標題"""
    click.echo()
    click.echo("=" * 70)
    click.echo(f"  {text}")
    click.echo("=" * 70)
    click.echo()


def print_success(text):
    """列印成功訊息"""
    click.secho(f"✅ {text}", fg='green')


def print_error(text):
    """列印錯誤訊息"""
    click.secho(f"❌ {text}", fg='red')


def print_info(text):
    """列印資訊"""
    click.secho(f"ℹ️  {text}", fg='blue')


def print_warning(text):
    """列印警告"""
    click.secho(f"⚠️  {text}", fg='yellow')


def read_excel_file(file_path):
    """
    讀取單個 Excel 檔案

    Returns:
        list: 單元格資料列表
    """
    cells_data = []

    try:
        # 打開檔案
        workbook = load_workbook(file_path, read_only=True, data_only=True)

        # 遍歷所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 遍歷所有單元格
            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is not None:
                        cells_data.append({
                            'sheet_name': sheet_name,
                            'row': row_idx,
                            'col': col_idx,
                            'location': f"{cell.column_letter}{cell.row}",
                            'value': str(cell.value),
                        })

        workbook.close()
        return cells_data

    except Exception as e:
        raise Exception(f"讀取檔案失敗: {e}")


# ============================================================================
# CLI 命令
# ============================================================================

@click.group()
@click.version_option(version='1.0.0', prog_name='excel-search')
def cli():
    """
    Excel 搜索系統 - 快速搜索大量 Excel 檔案

    使用方法：
        excel-search index <路徑>     # 索引檔案
        excel-search search <關鍵詞>  # 搜索
        excel-search stats            # 統計資訊
    """
    pass


@cli.command()
@click.argument('path', type=click.Path(exists=True))
@click.option('--recursive/--no-recursive', default=True, help='是否遞迴掃描子目錄')
def index(path, recursive):
    """
    索引 Excel 檔案

    PATH: 檔案或目錄路徑
    """
    print_header("📚 索引 Excel 檔案")

    # 初始化資料庫
    db = get_db()

    # 判斷是單個檔案還是目錄
    if os.path.isfile(path):
        files_to_index = [{
            'file_path': os.path.abspath(path),
            'file_name': os.path.basename(path),
            'file_size': os.path.getsize(path),
            'last_modified': datetime.fromtimestamp(os.path.getmtime(path))
        }]
        print_info(f"索引單個檔案: {os.path.basename(path)}")
    else:
        # 掃描目錄
        print_info(f"掃描目錄: {path}")
        scanner = FileScanner()
        files_to_index = scanner.scan_directory(path, recursive=recursive, show_progress=True)
        print_success(f"找到 {len(files_to_index)} 個 Excel 檔案")

    if not files_to_index:
        print_warning("沒有找到 Excel 檔案")
        return

    click.echo()

    # 索引檔案
    success_count = 0
    failed_count = 0
    total_cells = 0

    with tqdm(files_to_index, desc="索引中", unit="檔案") as pbar:
        for file_info in pbar:
            try:
                # 讀取檔案內容
                cells_data = read_excel_file(file_info['file_path'])

                if not cells_data:
                    pbar.write(f"⚠️  跳過空檔案: {file_info['file_name']}")
                    continue

                # 添加檔案記錄
                file_id = db.add_file(
                    file_path=file_info['file_path'],
                    file_name=file_info['file_name'],
                    last_modified=file_info['last_modified'],
                    file_size=file_info['file_size']
                )

                # 準備單元格資料
                cells_to_insert = []
                for cell in cells_data:
                    cells_to_insert.append({
                        'file_id': file_id,
                        'sheet_name': cell['sheet_name'],
                        'row': cell['row'],
                        'col': cell['col'],
                        'location': cell['location'],
                        'value': cell['value'],
                    })

                # 批量插入
                db.add_cells_batch(cells_to_insert)
                db.update_file_cell_count(file_id)

                success_count += 1
                total_cells += len(cells_data)
                pbar.set_postfix({'成功': success_count, '單元格': total_cells})

            except Exception as e:
                failed_count += 1
                pbar.write(f"❌ 索引失敗: {file_info['file_name']} - {e}")
                continue

    click.echo()
    print_header("📊 索引完成")

    print_success(f"成功索引: {success_count} 個檔案")
    if failed_count > 0:
        print_warning(f"失敗: {failed_count} 個檔案")
    print_info(f"總單元格數: {total_cells:,}")

    # 顯示統計
    stats = db.get_stats()
    print_info(f"資料庫大小: {stats['db_size_mb']} MB")

    db.close()


@cli.command()
@click.argument('keyword')
@click.option('--limit', default=20, help='最多顯示幾個結果')
@click.option('--full-row', is_flag=True, help='顯示完整行資料')
def search(keyword, limit, full_row):
    """
    搜索關鍵詞

    KEYWORD: 要搜索的關鍵詞
    """
    print_header(f"🔍 搜索: \"{keyword}\"")

    db = get_db()

    # 執行搜索
    cursor = db.conn.cursor()
    cursor.execute('''
        SELECT
            f.file_name,
            f.file_path,
            c.sheet_name,
            c.cell_location,
            c.value,
            c.row_num,
            c.col_num,
            c.file_id
        FROM cells c
        JOIN files f ON c.file_id = f.file_id
        WHERE c.value_lower LIKE ?
        ORDER BY f.file_name, c.sheet_name, c.row_num, c.col_num
        LIMIT ?
    ''', (f'%{keyword.lower()}%', limit))

    results = cursor.fetchall()

    if not results:
        print_warning(f"沒有找到包含 \"{keyword}\" 的結果")
        db.close()
        return

    print_success(f"找到 {len(results)} 個結果")
    click.echo()

    # 顯示結果
    for i, row in enumerate(results, 1):
        file_name, file_path, sheet_name, location, value, row_num, col_num, file_id = row

        click.echo("─" * 70)
        click.secho(f"結果 {i}", fg='cyan', bold=True)
        click.echo(f"📄 檔案: {file_name}")
        click.echo(f"📊 工作表: {sheet_name}")
        click.echo(f"📍 位置: {location} (第{row_num}行, 第{col_num}列)")

        # 高亮關鍵詞
        highlighted = value.replace(keyword, click.style(keyword, fg='yellow', bold=True))
        click.echo(f"📝 內容: {highlighted}")

        # 如果需要顯示完整行
        if full_row:
            cursor.execute('''
                SELECT cell_location, value
                FROM cells
                WHERE file_id = ? AND sheet_name = ? AND row_num = ?
                ORDER BY col_num
            ''', (file_id, sheet_name, row_num))

            row_cells = cursor.fetchall()
            if len(row_cells) > 1:
                click.echo(f"📋 完整行資料:")
                for cell_loc, cell_val in row_cells:
                    marker = " ← 匹配" if cell_loc == location else ""
                    click.echo(f"   {cell_loc:6s} = {str(cell_val)[:50]}{marker}")

        click.echo()

    if len(results) >= limit:
        print_info(f"僅顯示前 {limit} 個結果，使用 --limit 參數顯示更多")

    db.close()


@cli.command()
def stats():
    """顯示資料庫統計資訊"""
    print_header("📊 資料庫統計")

    db = get_db()
    stats = db.get_stats()

    click.echo(f"📁 索引檔案數:     {stats['file_count']:,}")
    click.echo(f"📝 總單元格數:     {stats['cell_count']:,}")
    click.echo(f"💾 檔案總大小:     {stats['total_file_size_mb']} MB")
    click.echo(f"🗄️  資料庫大小:     {stats['db_size_mb']} MB")
    click.echo(f"📊 平均單元格/檔案: {stats['avg_cells_per_file']:.0f}")
    click.echo(f"🕒 最後索引時間:   {stats['last_indexed'] or '尚未索引'}")

    click.echo()

    # 顯示最近索引的檔案
    cursor = db.conn.cursor()
    cursor.execute('''
        SELECT file_name, file_path, cell_count, indexed_at
        FROM files
        ORDER BY indexed_at DESC
        LIMIT 10
    ''')

    recent_files = cursor.fetchall()
    if recent_files:
        click.echo("📋 已索引的檔案清單:")
        click.echo()
        for i, (name, path, count, indexed) in enumerate(recent_files, 1):
            click.echo(f"  {i}. {name}")
            click.secho(f"     📍 路徑: {path}", fg='cyan')
            click.echo(f"     📊 單元格: {count} | 索引時間: {indexed}")
            click.echo()

    db.close()


@cli.command()
@click.confirmation_option(prompt='確定要清空資料庫嗎？')
def clear():
    """清空資料庫"""
    if os.path.exists(DATABASE_PATH):
        os.remove(DATABASE_PATH)
        print_success(f"已清空資料庫: {DATABASE_PATH}")
    else:
        print_info("資料庫不存在")


@cli.command()
def info():
    """顯示系統資訊"""
    print_header("ℹ️  系統資訊")

    click.echo(f"📁 資料庫路徑: {DATABASE_PATH}")
    click.echo(f"📏 資料庫大小: {format_size(os.path.getsize(DATABASE_PATH)) if os.path.exists(DATABASE_PATH) else '0 B'}")
    click.echo(f"🐍 Python 版本: {sys.version.split()[0]}")
    click.echo(f"📂 工作目錄: {os.getcwd()}")

    # 檢查資料庫是否存在
    if os.path.exists(DATABASE_PATH):
        print_success("資料庫已存在")
    else:
        print_warning("資料庫不存在，請先執行 index 命令")


# ============================================================================
# 主程式
# ============================================================================

if __name__ == '__main__':
    cli()
