"""
完整流程測試：讀取 Excel → 存入 SQLite → 搜索
"""
import sys
import os
from datetime import datetime
from openpyxl import load_workbook

# 引入我們的模組
from database import Database
from config import DATABASE_PATH

print("=" * 70)
print("🧪 Excel 搜索系統 - 完整流程測試")
print("=" * 70)
print()

# ============================================================================
# 步驟 1：選擇測試檔案
# ============================================================================
test_file = "Sharepoint/Hibiscus2 - Documents/01_Project_Management/1.08_Milestones/01_QG0/1170_State of the Art, documented/Start of the art evaluation_Hibiscus2.xlsx"

print(f"📄 測試檔案: {os.path.basename(test_file)}")
print(f"   大小: {os.path.getsize(test_file) / 1024:.2f} KB")
print()

# ============================================================================
# 步驟 2：讀取 Excel 檔案內容（簡化版 excel_reader）
# ============================================================================
print("📖 步驟 2：讀取 Excel 內容...")
print("-" * 70)

try:
    # 打開 Excel 檔案
    workbook = load_workbook(test_file, read_only=True, data_only=True)

    # 顯示工作表
    sheet_names = workbook.sheetnames
    print(f"✅ 成功打開檔案")
    print(f"   工作表數量: {len(sheet_names)}")
    print(f"   工作表名稱: {', '.join(sheet_names)}")
    print()

    # 讀取第一個工作表的內容
    sheet = workbook[sheet_names[0]]
    print(f"📊 讀取工作表: {sheet_names[0]}")
    print()

    # 收集所有單元格資料
    cells_data = []
    row_count = 0
    cell_count = 0

    print("   前 20 個單元格內容：")
    print("   " + "-" * 66)

    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        row_count = row_idx
        for col_idx, cell in enumerate(row, start=1):
            if cell.value is not None:
                cell_count += 1

                # 顯示前 20 個
                if cell_count <= 20:
                    location = f"{cell.column_letter}{cell.row}"
                    value = str(cell.value)[:50]  # 截斷太長的內容
                    print(f"   {location:6s} = {value}")

                # 收集資料
                cells_data.append({
                    'sheet_name': sheet_names[0],
                    'row': row_idx,
                    'col': col_idx,
                    'location': f"{cell.column_letter}{cell.row}",
                    'value': str(cell.value) if cell.value else ''
                })

    if cell_count > 20:
        print(f"   ... 還有 {cell_count - 20} 個單元格")

    print()
    print(f"✅ 讀取完成")
    print(f"   總行數: {row_count}")
    print(f"   有內容的單元格: {cell_count}")
    print()

    workbook.close()

except Exception as e:
    print(f"❌ 讀取失敗: {e}")
    sys.exit(1)

# ============================================================================
# 步驟 3：存入 SQLite 資料庫
# ============================================================================
print("💾 步驟 3：存入資料庫...")
print("-" * 70)

try:
    # 使用測試資料庫
    test_db_path = 'test_demo.db'
    if os.path.exists(test_db_path):
        os.remove(test_db_path)
        print(f"   刪除舊的測試資料庫")

    # 初始化資料庫
    db = Database(test_db_path)
    print(f"✅ 資料庫初始化完成: {test_db_path}")
    print()

    # 添加檔案記錄
    file_id = db.add_file(
        file_path=os.path.abspath(test_file),
        file_name=os.path.basename(test_file),
        last_modified=datetime.fromtimestamp(os.path.getmtime(test_file)),
        file_size=os.path.getsize(test_file)
    )
    print(f"✅ 檔案記錄已添加 (ID: {file_id})")

    # 準備單元格資料
    cells_to_insert = []
    for cell in cells_data:
        cells_to_insert.append({
            'file_id': file_id,
            'sheet_name': cell['sheet_name'],
            'row': cell['row'],
            'col': cell['col'],
            'location': cell['location'],
            'value': cell['value'],
        })

    # 批量插入
    db.add_cells_batch(cells_to_insert)
    print(f"✅ 單元格資料已插入: {len(cells_to_insert)} 個")

    # 更新檔案的單元格計數
    db.update_file_cell_count(file_id)

    # 顯示統計
    stats = db.get_stats()
    print()
    print(f"📊 資料庫統計:")
    print(f"   檔案數: {stats['file_count']}")
    print(f"   單元格數: {stats['cell_count']}")
    print(f"   資料庫大小: {stats['db_size_mb']} MB")
    print()

except Exception as e:
    print(f"❌ 存入資料庫失敗: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# ============================================================================
# 步驟 4：測試搜索功能
# ============================================================================
print("🔍 步驟 4：測試搜索功能...")
print("-" * 70)

# 找出一些可以搜索的關鍵詞
print("💡 讓我們找幾個關鍵詞來測試搜索...")
print()

# 取前幾個有值的單元格作為測試關鍵詞
test_keywords = []
for cell in cells_data[:10]:
    value = cell['value'].strip()
    if value and len(value) > 2:
        test_keywords.append(value)

if test_keywords:
    # 測試搜索第一個關鍵詞
    keyword = test_keywords[0]
    print(f"🔍 搜索關鍵詞: \"{keyword}\"")
    print()

    try:
        # 使用 SQL 搜索（部分匹配）
        cursor = db.conn.cursor()
        cursor.execute('''
            SELECT
                f.file_name,
                c.sheet_name,
                c.cell_location,
                c.value,
                c.row_num,
                c.col_num
            FROM cells c
            JOIN files f ON c.file_id = f.file_id
            WHERE c.value_lower LIKE ?
            LIMIT 10
        ''', (f'%{keyword.lower()}%',))

        results = cursor.fetchall()

        if results:
            print(f"✅ 找到 {len(results)} 個結果：")
            print()
            for i, row in enumerate(results, 1):
                print(f"   結果 {i}:")
                print(f"   📄 檔案: {row[0]}")
                print(f"   📊 工作表: {row[1]}")
                print(f"   📍 位置: {row[2]} (行{row[4]}, 列{row[5]})")
                print(f"   📝 內容: {row[3][:100]}")
                print()
        else:
            print(f"❌ 沒有找到包含 \"{keyword}\" 的結果")
            print()

    except Exception as e:
        print(f"❌ 搜索失敗: {e}")

    # 測試 FTS5 全文搜索
    print("-" * 70)
    print("🚀 測試 FTS5 全文搜索...")
    print()

    try:
        cursor.execute('''
            SELECT
                f.file_name,
                fts.sheet_name,
                fts.cell_location,
                fts.cell_value
            FROM content_fts fts
            JOIN files f ON fts.file_id = f.file_id
            WHERE content_fts MATCH ?
            LIMIT 5
        ''', (keyword,))

        results = cursor.fetchall()

        if results:
            print(f"✅ FTS5 找到 {len(results)} 個結果")
        else:
            print(f"ℹ️  FTS5 沒有找到結果（可能關鍵詞太短）")

    except Exception as e:
        print(f"⚠️  FTS5 搜索: {e}")

print()
print("=" * 70)
print("✅ 測試完成！")
print("=" * 70)
print()
print(f"📁 測試資料庫位置: {os.path.abspath(test_db_path)}")
print(f"💡 你可以用 SQLite 工具查看這個資料庫")
print()

# 關閉資料庫
db.close()
